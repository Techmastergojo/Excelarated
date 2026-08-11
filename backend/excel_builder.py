"""
Excel Builder: creates beautifully formatted .xlsx files from DataFrames.
Uses openpyxl for professional formatting: colored headers, alternating rows,
auto-width columns, and optionally embedded charts.
"""
import os
import re
from datetime import datetime
from typing import Dict, Any, List, Optional

import pandas as pd
import numpy as np

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

# ── Color palette ─────────────────────────────────────────
HEADER_BG   = "1A2942"   # Deep navy
HEADER_FG   = "00E5A0"   # Teal accent
ROW_ALT     = "F0F4F8"   # Light grey alt rows
ROW_NORMAL  = "FFFFFF"   # White rows
ACCENT1     = "00C48C"   # Teal
ACCENT2     = "7C3AED"   # Violet
BORDER_COL  = "D1D5DB"   # Subtle border


def _header_font():
    return Font(name="Calibri", bold=True, color=HEADER_FG, size=11)


def _data_font():
    return Font(name="Calibri", size=10)


def _header_fill():
    return PatternFill("solid", fgColor=HEADER_BG)


def _alt_fill():
    return PatternFill("solid", fgColor=ROW_ALT)


def _border():
    side = Side(style="thin", color=BORDER_COL)
    return Border(left=side, right=side, top=side, bottom=side)


def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)


def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_dataframe(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1):
    """Write a DataFrame to a worksheet with full formatting."""
    # Headers
    for ci, col in enumerate(df.columns, start=start_col):
        cell = ws.cell(row=start_row, column=ci, value=str(col))
        cell.font      = _header_font()
        cell.fill      = _header_fill()
        cell.alignment = _center()
        cell.border    = _border()

    # Data rows
    for ri, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        fill = _alt_fill() if ri % 2 == 0 else PatternFill("solid", fgColor=ROW_NORMAL)
        for ci, col in enumerate(df.columns, start=start_col):
            val = row[col]
            # Convert numpy types to Python native
            if isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = float(val) if not np.isnan(val) else None
            elif pd.isna(val):
                val = None

            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = _data_font()
            cell.fill      = fill
            cell.border    = _border()
            cell.alignment = _left()

    # Auto-fit column widths
    for ci, col in enumerate(df.columns, start=start_col):
        col_letter = get_column_letter(ci)
        max_len = max(
            len(str(col)),
            *[len(str(df.iloc[ri][col])[:60]) for ri in range(min(len(df), 100))]
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)

    # Freeze header row
    ws.freeze_panes = ws.cell(row=start_row + 1, column=start_col)


def _add_chart(wb: Workbook, ws, df: pd.DataFrame, chart_type: str = "bar",
               x_col: str = None, y_col: str = None, title: str = "Chart",
               anchor: str = "A1"):
    """Add an openpyxl chart to a sheet."""
    numeric_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
    text_cols    = [c for c in df.columns if not pd.api.types.is_numeric_dtype(df[c])]

    x_col = x_col or (text_cols[0] if text_cols else df.columns[0])
    y_col = y_col or (numeric_cols[0] if numeric_cols else df.columns[-1])

    if y_col not in df.columns or x_col not in df.columns:
        return

    # Build a temporary chart sheet
    chart_ws = wb.create_sheet(f"{title[:20]}_Chart")
    chart_ws["A1"] = x_col
    chart_ws["B1"] = y_col
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        chart_ws.cell(row=i, column=1, value=str(row[x_col]))
        val = row[y_col]
        if pd.notna(val):
            chart_ws.cell(row=i, column=2, value=float(val))

    rows_written = len(df) + 1

    cats = Reference(chart_ws, min_col=1, min_row=2, max_row=rows_written)
    data = Reference(chart_ws, min_col=2, min_row=1, max_row=rows_written)

    if chart_type == "line":
        chart = LineChart()
    elif chart_type == "pie":
        chart = PieChart()
    else:
        chart = BarChart()

    chart.add_data(data, titles_from_data=True)
    if chart_type != "pie":
        chart.set_categories(cats)

    chart.title  = title
    chart.width  = 20
    chart.height = 12

    # Style
    if hasattr(chart, "shape"):
        chart.shape = 4  # Rounded

    ws.add_chart(chart, anchor)


EXPORTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "exports")
os.makedirs(EXPORTS_DIR, exist_ok=True)


def build_excel(
    dfs: Dict[str, pd.DataFrame],
    output_name: str,
    title: str = "Excelarated Report",
    add_charts: bool = True,
    chart_configs: List[Dict] = None,
    summary_text: str = None,
) -> str:
    """
    Build a beautifully formatted Excel file.
    
    dfs: dict of {sheet_name: DataFrame}
    Returns the absolute path to the saved file.
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default empty sheet

    # Cover / Summary sheet
    cover = wb.create_sheet("Summary", 0)
    cover.sheet_view.showGridLines = False

    # Title cell
    cover["B2"] = title
    cover["B2"].font = Font(name="Calibri", size=20, bold=True, color=HEADER_BG)

    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    cover["B3"] = f"Generated by Excelarated  •  {ts}"
    cover["B3"].font = Font(name="Calibri", size=11, color="6B7280")

    if summary_text:
        cover["B5"] = summary_text
        cover["B5"].font = Font(name="Calibri", size=11, color="374151")
        cover["B5"].alignment = Alignment(wrap_text=True)
        cover.column_dimensions["B"].width = 70
        cover.row_dimensions[5].height = 60

    row_offset = 7
    cover["B" + str(row_offset)] = "Contents"
    cover["B" + str(row_offset)].font = Font(name="Calibri", bold=True, size=12, color=HEADER_BG)

    for i, sheet_name in enumerate(dfs.keys()):
        df = dfs[sheet_name]
        cover[f"B{row_offset + 1 + i}"] = f"  {i+1}. {sheet_name} — {len(df):,} rows, {len(df.columns)} columns"
        cover[f"B{row_offset + 1 + i}"].font = Font(name="Calibri", size=10, color="374151")

    cover.column_dimensions["A"].width = 3
    cover.column_dimensions["B"].width = 65

    # Data sheets
    for sheet_name, df in dfs.items():
        ws = wb.create_sheet(sheet_name[:31])  # Excel sheet name max 31 chars
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 22

        _write_dataframe(ws, df, start_row=1, start_col=1)

        # Add chart if requested
        if add_charts and len(df) > 0:
            numeric_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
            if numeric_cols:
                anchor_col = get_column_letter(len(df.columns) + 2)
                chart_cfg  = {}
                if chart_configs:
                    chart_cfg = next((c for c in chart_configs if c.get("sheet") == sheet_name), {})

                _add_chart(
                    wb, ws, df.head(30),
                    chart_type=chart_cfg.get("type", "bar"),
                    x_col=chart_cfg.get("x_col"),
                    y_col=chart_cfg.get("y_col"),
                    title=f"{sheet_name} Chart",
                    anchor=f"{anchor_col}2",
                )

    # Save
    safe_name = re.sub(r"[^\w\-.]", "_", output_name)
    if not safe_name.endswith(".xlsx"):
        safe_name += ".xlsx"
    out_path = os.path.join(EXPORTS_DIR, safe_name)
    wb.save(out_path)
    return out_path
